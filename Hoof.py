# import csv
# import pandas as pd

# with open('newFile.CSV', 'r') as f:
#    reader = csv.reader(f)
#    for row in reader:
#        print (row)

# f = pd.read_csv("newFile.CSV")
# keep_col = ['Subject', 'From: (Name)', 'To: (Name)', 'CC: (Name)', 'Categories']
# new_f = f[keep_col]
# new_f.to_csv("newFile.csv", index=False)
# print(new_f)


# integers = number
# float = number.0
# string/chart = Hamza
# boolean = true or false

# print('what is your name, and what is your favorite color ?')
# name = input('Hi, My name is ')
# color= input('my favorite color is ')
# print (name + ' likes ' + color)


# birth_year = input('Birth year: ')
# print(type(birth_year))
# age = 2019 - int(birth_year)
# print(type(age))
# print(age)

# weight = input('insert your weight in (lbs) ')
# weight = float(weight) / 2.2046
# print(str(weight) + ' kg')

# weight_lbs = input('wight (lbs): ')
# weight_kg = float(weight_lbs) * 0.45
# print("your weight is" + str(weight_kg))

# STRINGS
# course ='Python for Beginners'
# another = course[:]
# print(another[:])

# name = 'Jennifer'
# print(name[1:-1])

# Formatted Strings
# first = 'John'
# last = 'Smith'
# message = first + ' [' + last + '] is a coder'
# msg = f'{first} [{last}] is a coder'
# print(msg)

#
# course = 'Python for Beginners'
# print(len(course))
# print(course.upper())
# print(course.lower())
# print(course.replace('Beginners', 'Absolute begginers'))
# print('Python' in course)

# rekap len() count lenghth upper and lower and title and find replace and in operatore

# Arithmetic Operations

# print(10 - 3 + 14 - ((19*21)/3) // 3)

# Incrementation
# x = 10
# x -= 3
# print(x)

# Operator precedence

# x = 10 + 3 * 2
# print(x)
# exponentionation
# multiplication or division
# addition or substraction

# x = (2 + 3) * 10 -3
# print(x)

# Math Functions

# import math

# print(math.floor(2.9))
# x = -2.9
# print(round(x))
# print(abs(x)) #always positive number (absolute)

# if statement
# is_hot = False
# is_cold = True

# if is_hot:
#    print("it's a hot day")
#    print("Drink plenty of water")
# elif is_cold:
#    print("it's a cold day")
#    print("Wear warm clothes")
# else:
#    print("it's a lovely day")
# print("Enjoy your day")

#
'''
good_credit = False
price = 100000
if good_credit:
    print("you have a good credit")
    print("you only need to put down 10% as a down payment which will be " + str(price*0.1) + "$")
else:
    print("you have a bad credit")
    print("you have to put down 20% of down payment which will be " + str(price*0.2) + "$")
'''

'''
#logical operators OR AND NOT < > = >= 
has_high_income = True
has_good_credit = False
criminal_record = True
if has_good_credit or has_high_income:
    print("Eligible for loan")
else:
    print("NOT eligible for loan")

'''
'''
name_lengh = input("Please enter your name ")
if len(name_lengh) < 3:
    print("name must be at least 3 characters")
elif len(name_lengh) > 6:
    print(" name can be a maximum of 6 characters")
else:
    print('name looks good')



print("Welcome to your weight calculator")
weight = int(input('Please enter your weight: '))
unit = input("(L)bs or (K)g: ")
if unit.upper() =="l":
    converted = weight * 0.45
    print(f"You are {converted} kilos")
else:
    converted = weight / 0.45
    print(f"You are {converted} pounds")



#Loops

i = 1
while i <= 5:
    print('*' * i)
    i = i + 1
print("Done")

#guessing Game
secret_number = 9
i = 0 #guess count
while i < 3: #guess_limit:
    guess =int(input('Guess: '))
    i +=1
    if guess == secret_number:
        print('You won')
        break
else:
    print('You failed')


#Car Game

welcome = print('welcome to the Car Game')
key_word = input('Please type Start, Stop or Quit: ')
if key_word.lower() == 'start':
    print('Car started...ready to go! ')
elif key_word.lower() == 'stop':
    print('Car stopped. ')
elif key_word.lower() == 'quit':
    print('Quitting game')
else :
    print("I don't understand")



command = ""
started = False
while True:
    command = input("> ").lower()
    if command== "start":
        if started:
           print("Car has already started!")
        else:
           started = True
           print ("Car started...")
    elif command == "stop":
        if not started:
           print("Car has already stopped!!!")
        else:
           started = False
           print("Car stopped.")
    elif command == "help":
        print ("""
start - to start the car
stop - to stop the car
quit - to quit
        """)
    elif command == "quit":
        break
    else:
        print ("sorry i don't understand the word")


#For Loops used for itterations

prices = [10, 20, 30]
total = 0
for price in prices:
    total += price
    print(f"Total : {total}")


#Nested Loops means adding a loop inside another loop


(0, 0)
(0, 1)
(0, 2)
(1, 0)
(1, 1)
(1, 2)

for x in range(4):
    for y in range(3):
        print(f'({x}, {y})')

numbers = [5, 2, 5, 2, 2]
for i in numbers:
    result = ''
    for x in range(i):
        result += 'x'
    print(result)



#Lists

names = ['hamza', 'Rita', 'Yeliz', 'Mads', 'Sveta']
names[0] = 'Hamzah'
print(names)

#Find the largest number in a list

numbers = [1, 2, 3, 4, 5, 6, 7, 8, 200]
max = numbers[0]
for item in numbers:
    if item > max:
        max = item
print(max)



# 2D lists

matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
for row in matrix:
    for item in row:
        print(item)


#List methods operations on lists

numbers = [5, 2, 1, 7, 4]
number2 = numbers.copy()
numbers.append(12)
print(numbers)
print(number2)

numbers =[5, 2, 3, 2, 3, 1, 5, 3]
uniques =[]
for items in numbers:
    if items not in uniques:
        uniques.append(items)
print(uniques)

#Tuple


#Unpacking
coordinates = (1, 2, 3)
x, y, z = coordinates
print(y)


#Dictionries


customer = {
    "name" : "John Smith",
    "age": "30",
    "is verified": True
}
customer["birthdays"] = "Jan 01 1980"
print(customer.get("birthdays"))

phone = input("Phone: ")
digits_mapping = {
    "1": "One",
    "2": "two",
    "3": "three",
    "4": "four",
}
output = ""
for ch in phone:
    output += digits_mapping.get(ch, " Does Not Exist ") + " "
print(output)



message = input (">")
words = message.split(' ')
emojis = {
    ":)": ""

}
output =""
for word in words:
    output += emojis.get(word, word) +  " "
print(output)


#Functions

def greet_user():
    print('Hi there!')
    print('Welcome abroad')


print("start")
greet_user()
print("Finish")


#Parameters

def greet_user(first_name, last_name):
    print(f'Hi {first_name} {last_name}!')
    print('Welcome abroad')


print("start")
greet_user("John", last_name="smith")
greet_user("Mary", last_name="blonde")
print("Finish")


# Return statement

def square(number):
    return number * number


print(square(3))


#create a reusable function

def emoji_converter(message):
    words = message.split(" ")
    emojis = {
        ":)": "lol",
        ":(": "MDR"
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return output


message = input(">")
print(emoji_converter(message))

# Exceptions
try:
    age = int(input('Age: '))
    income = 20000
    risk = income / age
    print(age)
except ZeroDivisionError:
    print('Age can not be Zero')
except ValueError:
    print('Invalid Value')


#Classes

class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y
    def move(self):
        print("move")

    def draw(self):
        print("draw")

#point1 = Point()
#point1.x = 10
#point1.y = 20
#print(point1.x)
#point1.draw()

#Constractors

point = Point(10, 20)
print(point.x)

class Person:
    def __init__(self, name):
        self.name = name

    def talk(self):
        print(f"Hi, I am {self.name}")


john = Person("John Smith")
john.talk()

bob = Person("Bob Smith")
bob.talk()



class Mammal:
    def walk(self):
        print("walk")
class Dog(Mammal):
    def bark(self):
        print("bark")

class Cat(Mammal):
    def be_annoying(self):
        print("anoying")

dog1 = Dog()
dog1.bark()
cat1 = Cat()
cat1.be_annoying()



#Modules

def lbs_to_kg(weight):
    return weight * 0.45


def kg_to_lbs(weight):
    return wight / 0.45



from converter import kg_to_lbs

print(kg_to_lbs(100))


numbers = [120, 21, 23, 20]
print(max(numbers))



#Package

#Generating Random Values

import random

members = ['SVEV', 'BARN', 'RGRY', 'ASKO', 'TXAS', 'POOM', 'AHCD']
leader = random.choice(members)
print(leader)


class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first, second


dice = Dice()
print(dice.roll())


# Files and Directories

# Excel Spreadsheet

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename) #only relying on the mentioned path
    #if the above code/path doesn't work put "r" in front of the path
    sheet = wb['Sheet1']
    #this is just to show that the code works cell = sheet['a1']
    #cell = sheet.cell(1, 1)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row = 2,
              max_row = sheet.max_row,
              min_col = 4,
              max_col = 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'm2') #sometimes ther's an issue with naming column and rows and file should be closed before update
    wb.save(filename)

'''

# Machine Learning

'''
1- Import
2- Clean
Splitinto training and tests
4- Create Modeæs (Decision trees, algorithms)
5- Train model
6- Prediction
7- Evaluate


from sklearn import datasets

iris = datasets.load_iris()
digits = datasets.load_digits()
print(digits.data)
digits.images[0]

'''

import collections
import pandas as pd

# %matplotlib inline

file = open('newFile.CSV', "r", encoding='UTF8')
a = file.read()
stopwords = set(line.strip() for line in open ('stopwords.txt'))
stopwords = stopwords.union(set(['qam','barn', 'ahcd', 'svev','rgry','addz','ditw','txas', 'mcni', 'pimn', 'poom','asko']))

wordcount = {}

ignore = {'change','Diabetes'}
for word in a.lower().split():
    word = word.replace(".","")
    word = word.replace(",","")
    word = word.replace(":","")
    word = word.replace("\"","")
    word = word.replace("!","")
    word = word.replace("â€œ","")
    word = word.replace("â€˜","")
    word = word.replace("*","")
    if word not in stopwords:
        if word not in wordcount:
            wordcount[word] = 1
        else:
            wordcount[word] += 1
# Print most common word
n_print = int(input("Who is frequently contacting us ? : "))
print("\nOK. QAM's {} most common customers are: \n".format(n_print))
word_counter = collections.Counter(wordcount)
for word, count in word_counter.most_common(n_print):
    print(word, ": ", count)
# Close the file
file.close()

file = open('Subjects.CSV', "r", encoding='UTF8')
a = file.read()
stopwords = set(line.strip() for line in open ('stopwords.txt'))
stopwords = stopwords.union(set(['qam','barn', 'ahcd', 'svev','rgry','addz','ditw','txas', 'mcni', 'pimn', 'poom','asko', 'quality']))

wordcount = {}

ignore = {'change','Diabetes'}
for word in a.lower().split():
    word = word.replace(".","")
    word = word.replace(",","")
    word = word.replace(":","")
    word = word.replace("\"","")
    word = word.replace("!","")
    word = word.replace("â€œ","")
    word = word.replace("â€˜","")
    word = word.replace("*","")
    if word not in stopwords:
        if word not in wordcount:
            wordcount[word] = 1
        else:
            wordcount[word] += 1
# Print most common word
n_print = int(input("\nHow many most common words to print :  "))
print("\nOK. The {} most common words in subjects are\n".format(n_print))
word_counter = collections.Counter(wordcount)
for word, count in word_counter.most_common(n_print):
    print(word, ": ", count)
# Close the file
file.close()

file = open('Discussion.CSV', "r", encoding='UTF8')
a = file.read()
stopwords = set(line.strip() for line in open('stopwords.txt'))
stopwords = stopwords.union(set(['qam', 'barn', 'ahcd', 'svev', 'rgry', 'addz', 'ditw', 'txas', 'mcni', 'pimn', 'poom', 'asko', 'quality', 'to','from', 'sent', 'a', 'an', 'is']))

wordcount = {}

ignore = {'change', 'Diabetes'}
for word in a.lower().split():
    word = word.replace(".", "")
    word = word.replace(",", "")
    word = word.replace(":", "")
    word = word.replace("\"", "")
    word = word.replace("!", "")
    word = word.replace("â€œ", "")
    word = word.replace("â€˜", "")
    word = word.replace("*", "")
    if word not in stopwords:
        if word not in wordcount:
            wordcount[word] = 1
        else:
            wordcount[word] += 1
# Print most common word
n_print = int(input("\nHow many most common words to print :  "))
print("\nOK. The {} most common words in communications are : \n".format(n_print))
word_counter = collections.Counter(wordcount)
for word, count in word_counter.most_common(n_print):
    print(word, ": ", count)
# Close the file
file.close()

'''
# Create a data frame of the most common words

# Draw a bar chart
lst = word_counter.most_common(n_print)
df = pd.DataFrame(lst, columns = ['Word', 'Count'])
df.plot.bar(x='Word' , y='Count', rot=0)

file = open('/Users/addz/PycharmProjects/QAM/newFile.CSV',"r+")
a = file.read()
stopwords = set(line.strip() for line in open ('stopwords.txt'))
stopwords = stopwords.union(set(['change', 'at', 'actions','and','with','of', 'by', '&', 'no', 'business', 'area', 'region', 'as','international','operations','novo', 'nordisk','a/s']))

wordcount = {}

ignore = {'change','the','a','if','in','it','of','or', 'A/S', 'Novo', 'Nordisk', 'Regulatory', 'local', 'Quality', 'Dept.','FLIT', 'Dept.', 'Operations', 'Diabetes', 'novo', 'nordisk', 'operations', 'a/s'}
for word in a.lower().split():
    word = word.replace(".","")
    word = word.replace(",","")
    word = word.replace(":","")
    word = word.replace("\"","")
    word = word.replace("!","")
    word = word.replace("/", "")
    word = word.replace("â€œ","")
    word = word.replace("â€˜","")
    word = word.replace("*","")
    if word not in stopwords:
        if word not in wordcount:
            wordcount[word] = 1
        else:
            wordcount[word] += 1
# Print most common word
n_print = int(input("How many most common words to print: "))
print("\nOK. The {} most common words are as follows\n".format(n_print))
word_counter = collections.Counter(wordcount)
for word, count in word_counter.most_common(n_print):
    print(word, ": ", count)
# Close the file
file.close()
# Create a data frame of the most common words
# Draw a bar chart
lst = word_counter.most_common(n_print)
df = pd.DataFrame(lst, columns = ['Word', 'Count'])
df.plot.bar(x='Word',y='Count')




file = open("STAFF.CSV","r+")
a = file.read()
stopwords = set(line.strip() for line in open ('stopwords.txt'))
stopwords = stopwords.union(set(['change', 'at', 'actions','and','with','of','mm', 'you', 'is', 'this', 'group', 'by', '&', 'no', 'business', 'area', 'region', 'as','international','operations','novo', 'nordisk','a/s','that','subject','sent']))

wordcount = {}

ignore = {'change','the','a','if','in','it','of','or', 'A/S', 'Novo', 'Nordisk', 'Regulatory', 'Dept.', 'FLIT', 'Dept.', 'Operations', 'Diabetes', 'novo', 'nordisk', 'operations', 'a/s'}
for word in a.lower().split():
    word = word.replace(".","")
    word = word.replace(",","")
    word = word.replace(":","")
    word = word.replace("\"","")
    word = word.replace("!","")
    word = word.replace("/", "")
    word = word.replace("â€œ","")
    word = word.replace("â€˜","")
    word = word.replace("*","")
    if word not in stopwords:
        if word not in wordcount:
            wordcount[word] = 1
        else:
            wordcount[word] += 1
# Print most common word
n_print = int(input("How many most common words to print: "))
print("\nOK. The {} most common words are as follows\n".format(n_print))
word_counter = collections.Counter(wordcount)
for word, count in word_counter.most_common(n_print):
    print(word, ": ", count)
# Close the file
file.close()
# Create a data frame of the most common words
# Draw a bar chart
lst = word_counter.most_common(n_print)
df = pd.DataFrame(lst, columns = ['Word', 'Count'])
df.plot.bar(x='Word',y='Count')
'''
'''
import csv

a = "Validation"  # String that you want to search
with open('Discussion.CSV') as f_obj:
    reader = csv.reader(f_obj, delimiter=',')
    for line in reader:  # Iterates through the rows of your csv
        print(line)  # line here refers to a row in the csv
        if a in line:  # If the string you want to search is in the row
            print("String found in first row of csv")

'''
